#! python3
# emailNotifier.py

import openpyxl, smtplib, sys, getpass
from twilio.rest import Client

account_sid = 'AC4fc72ca66dea1fee1f3cbad74ad8f443'
auth_token = '90b951e679f729294a9df04fcdf5a4c9'
client = Client(account_sid, auth_token)




#open the spreadsheet to get values

wb = openpyxl.load_workbook('testemailer.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastRow = sheet.max_row


nestedDict = {}



for r in range(2, lastRow+1):
	nameTester = sheet.cell(row = r, column=1).value
	nestedDict[nameTester] = {}
	nestedDict[nameTester]['email'] = sheet.cell(row = r, column=2).value
	nestedDict[nameTester]['case'] = sheet.cell(row = r, column=3).value
	nestedDict[nameTester]['location'] = sheet.cell(row = r, column=4).value
	nestedDict[nameTester]['timeStart'] = sheet.cell(row = r, column=5).value
	nestedDict[nameTester]['phoneNumb'] = sheet.cell(row = r, column=6).value



# log into email

print('Input your email password')
password = getpass.getpass()

smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('william.corrigan@epiqsystems.com', password)



# send out emails

for d_id, d_info in nestedDict.items():
	body = "Subject: Case Tomorrow. \n\n Dear %s, \n\n Case name: %s \n\n Location: %s \n\n Time: %s. \n\n Kind regards, \n\n Will Corrigan" %(d_id, nestedDict[d_id]['case'], nestedDict[d_id]['location'], nestedDict[d_id]['timeStart'])
	print('Sending email to %s...' % nestedDict[d_id]['email'])
	sendmailStatus = smtpObj.sendmail('william.corrigan@epiqsystems.com', nestedDict[d_id]['email'], body.encode('utf-8'))

	message = client.messages \
	                .create(
	                     body=body,
	                     from_='+441452204141',
	                     to=nestedDict[d_id]['phoneNumb']
	                 )



	if sendmailStatus != {}:
		print('There was a problem sending email to %s: %s' %(nestedDict[d_id]['email'], sendmailStatus))


smtpObj.quit()
